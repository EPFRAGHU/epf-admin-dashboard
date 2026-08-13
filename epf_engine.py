"""
epf_engine.py
-------------
Core data model, contribution calculations, and Excel (Form 3A / Form 6A)
generation for the EPF Data Entry program.

This module has NO GUI code in it, so it can be tested / reused on its own.
"""

import json
import os
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from typing import List

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break


def normalize_member_id(mid):
    s = str(mid).strip()
    if s.startswith("__UAN__"):
        return s
    return s[-7:] if len(s) >= 7 else s

def format_member_id_display(mid):
    s = str(mid).strip()
    if s.startswith("__UAN__"):
        return ""
    return s

def natural_sort_key(text: str):
    """
    Sort key that treats runs of digits as numbers, so account numbers like
    "1", "2", "10", "11" sort as 1, 2, 10, 11 -- not as plain text
    ("1", "10", "11", "2", ...). Works fine on mixed values too, e.g.
    "OR/2", "OR/10" -> sorts by the non-digit parts, then the numeric parts.
    """
    return [int(chunk) if chunk.isdigit() else chunk.lower()
            for chunk in re.split(r"(\d+)", text or "")]


# Standard reasons an employee can leave an establishment (used by the
# Employee Master "Reason of Leaving" field).
REASONS_FOR_LEAVING = ["", "Cessation", "Resigned", "Superannuation",
                        "Death in Service", "Death away from Service"]

SUPERANNUATION_AGE = 58  # EPF Scheme retirement/superannuation age


def calc_age_years(dob_text: str, as_of: date = None):
    """
    Returns whole years of age given a DD/MM/YYYY date-of-birth string, or
    None if dob_text is blank/unparseable. `as_of` defaults to today.
    """
    dob_text = (dob_text or "").strip()
    if not dob_text:
        return None
    try:
        dob = datetime.strptime(dob_text, "%d/%m/%Y").date()
    except ValueError:
        return None
    as_of = as_of or date.today()
    years = as_of.year - dob.year - ((as_of.month, as_of.day) < (dob.month, dob.day))
    return years

MONTHS = [
    "Mar Paid in Apr", "Apr Paid in May", "May Paid in Jun", 
    "Jun Paid in Jul", "Jul Paid in Aug", "Aug Paid in Sep", 
    "Sep Paid in Oct", "Oct Paid in Nov", "Nov Paid in Dec", 
    "Dec Paid in Jan", "Jan Paid in Feb", "Feb Paid in Mar"
]
MONTH_FULL = {
    "APR": "April", "MAY": "May", "JUN": "June", "JUL": "July", "AUG": "August",
    "SEP": "September", "OCT": "October", "NOV": "November", "DEC": "December",
    "JAN": "January", "FEB": "February", "MAR": "March",
}

def get_wage_ceiling(month_idx: int, year_from: str) -> float:
    try:
        y_from = int(year_from)
    except:
        return 15000.0
    
    y = y_from if month_idx < 9 else y_from + 1
    m = month_idx + 4 if month_idx < 9 else month_idx - 8
    
    ym = y * 100 + m
    
    if ym >= 201409: return 15000.0
    if ym >= 200106: return 6500.0
    if ym >= 199410: return 5000.0
    if ym >= 199011: return 3500.0
    if ym >= 198509: return 2500.0
    if ym >= 197612: return 1600.0
    if ym >= 196212: return 1000.0
    if ym >= 195706: return 500.0
    return 300.0

def get_wage_ceilings_for_year(year_from: str) -> List[float]:
    return [get_wage_ceiling(i, year_from) for i in range(12)]

# --------------------------------------------------------------------------
# Contribution schemes
# --------------------------------------------------------------------------
# PRE_1997  -- (up to currency year 1996-97): EPF and FPF rates are the SAME
#              for worker and employer (e.g. EPF 6.84% + FPF/EPS 1.16% = 8.33%
#              from either side, or whatever rate applies to that year).
# POST_1997 -- (1997-98 onwards): worker's entire 12% goes to EPF (Account 1).
#              Employer's 12% is split into Pension Fund/EPS 8.33%
#              (Account 10) + EPF 3.67% (Account 1).
SCHEME_PRE_1997 = "pre_1997"
SCHEME_POST_1997 = "post_1997"

# Header-keyword -> month index, used by the bulk Excel importer below.
# Order matters only in that each key is a distinct 3-letter code, so there's
# no ambiguity when matching lower-cased header text with "in".
_MONTH_HEADER_KEYS = {
    "mar": 0, "apr": 1, "may": 2, "jun": 3, "jul": 4, "aug": 5,
    "sep": 6, "oct": 7, "nov": 8, "dec": 9, "jan": 10, "feb": 11,
}


def get_excel_sheet_names(filepath: str):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    return wb.sheetnames

def import_wages_from_excel(filepath: str, sheet_name=None, import_type="yearly", month_idx=None):
    """
    Reads a flat Excel sheet with columns such as:
        SL | Member ID | Name | APR | MAY | ... | MAR | Total Wages
    (column order/wording can vary a little -- headers are matched by
    keyword, case-insensitively).

    Returns (records, warnings):
        records  -- list of dicts: {"member_id": str, "name": str, "wages": [12 floats]}
        warnings -- list of human-readable strings describing any rows/values
                    that needed to be skipped or defaulted, so the caller can
                    show them to the user.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Find the header row: scan the first several rows for one containing
    # a cell that looks like an "Member ID" heading.
    header_row_idx = None
    header_cells = []
    for r in range(1, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        row_text = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        has_id = any(("account" in v or "member id" in v or "uan" in v) for v in row_text)
        has_name = any(("name" in v) for v in row_text)
        if has_id and has_name:
            header_row_idx = r
            header_cells = row_text
            break

    if header_row_idx is None:
        raise ValueError("Could not find a header row containing 'Member ID' or 'UAN' in the sheet.")

    col_map = {"months": {}, "gross_months": {}, "ncp_months": {}}
    for idx, text in enumerate(header_cells):
        if not text:
            continue
        if "uan" in text:
            col_map["uan"] = idx
        elif "member id" in text:
            col_map["member_id"] = idx
        elif "account" in text and "bank" not in text and "member_id" not in col_map:
            col_map["member_id"] = idx
        elif "name" in text and "father" not in text and "husband" not in text:
            col_map.setdefault("name", idx)
        elif "father" in text or "husband" in text: col_map["father_name"] = idx
        elif "birth" in text or "dob" in text: col_map["dob"] = idx
        elif "sex" in text or "gender" in text: col_map["sex"] = idx
        elif "joining" in text or "doj" in text: col_map["doj"] = idx
        elif "exit" in text or "leaving" in text and "reason" not in text: col_map["doe"] = idx
        elif "reason" in text: col_map["reason_leaving"] = idx
        elif text == "sl" or text == "sl no": col_map["serial_no"] = idx
        elif "higher" in text and "epf" in text: col_map["higher_epf"] = idx
        elif "age" in text and "58" in text: col_map["age_crosses_58"] = idx
        elif text == "total":
            col_map["total"] = idx
        else:
            if import_type == "monthly":
                if "gross" in text:
                    col_map["gross_wages_single"] = idx
                elif "ncp" in text:
                    col_map["ncp_single"] = idx
                elif "epf" in text or "wage" in text:
                    col_map["epf_wages_single"] = idx
            else:
                for key, m_idx in _MONTH_HEADER_KEYS.items():
                    if key in text:
                        if "gross" in text:
                            col_map["gross_months"][m_idx] = idx
                        elif "ncp" in text:
                            col_map["ncp_months"][m_idx] = idx
                        else:
                            col_map["months"][m_idx] = idx
                        break

    if "uan" not in col_map or "name" not in col_map:
        raise ValueError("Could not find both 'UAN' and a 'Name' column in the header row. Please check the column headings.")

    def cell_val(row_cells, idx):
        if idx is None or idx >= len(row_cells):
            return None
        return row_cells[idx].value

    records = []
    warnings = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = ws[r]
        uan_val = cell_val(row_cells, col_map.get("uan"))
        if uan_val is None or str(uan_val).strip() == "":
            continue
        acc = f"__UAN__{str(uan_val).strip()}"
        name_val = cell_val(row_cells, col_map.get("name"))
        name = str(name_val).strip() if name_val is not None else ""
        if not name:
            warnings.append(f"Row {r}: no employee name given for account {acc} "
                             f"(you can fill this in later via the Employee Master).")

        wages = [0.0] * 12
        gross_wages = [0.0] * 12
        ncp_days = [0] * 12
        
        if import_type == "monthly" and month_idx is not None and 0 <= month_idx < 12:
            g_val = cell_val(row_cells, col_map.get("gross_wages_single"))
            w_val = cell_val(row_cells, col_map.get("epf_wages_single"))
            n_val = cell_val(row_cells, col_map.get("ncp_single"))
            
            try:
                gross_wages[month_idx] = float(g_val) if g_val is not None and g_val != "" else 0.0
            except (TypeError, ValueError):
                pass
            
            try:
                wages[month_idx] = float(w_val) if w_val is not None and w_val != "" else 0.0
            except (TypeError, ValueError):
                warnings.append(f"Row {r}: could not read wage value for account {acc} ('{w_val}') -- treated as 0.")

            try:
                ncp_days[month_idx] = int(n_val) if n_val is not None and n_val != "" else 0
            except (TypeError, ValueError):
                pass
        else:
            for m_idx in range(12):
                col_idx = col_map["months"].get(m_idx)
                val = cell_val(row_cells, col_idx)
                if val is None or val == "":
                    wages[m_idx] = 0.0
                else:
                    try:
                        wages[m_idx] = float(val)
                    except (TypeError, ValueError):
                        wages[m_idx] = 0.0
                        warnings.append(f"Row {r}: could not read the {MONTHS[m_idx]} wage value "
                                         f"for account {acc} ('{val}') -- treated as 0.")
                
                g_col_idx = col_map["gross_months"].get(m_idx)
                g_val = cell_val(row_cells, g_col_idx)
                if g_val is None or g_val == "":
                    gross_wages[m_idx] = wages[m_idx]
                else:
                    try:
                        gross_wages[m_idx] = float(g_val)
                    except (TypeError, ValueError):
                        gross_wages[m_idx] = wages[m_idx]
                        
                n_col_idx = col_map["ncp_months"].get(m_idx)
                n_val = cell_val(row_cells, n_col_idx)
                if n_val is not None and n_val != "":
                    try:
                        ncp_days[m_idx] = int(n_val)
                    except (TypeError, ValueError):
                        pass

        def format_date(val):
            import pandas as pd
            if pd.isna(val) or val == "": return ""
            if isinstance(val, str):
                val = val.strip().replace("/", "-")
                if val.endswith(" 00:00:00"): val = val.split(" ")[0]
                return val
            if hasattr(val, "strftime"):
                return val.strftime("%d-%m-%Y")
            return str(val)

        def get_val(key):
            val = cell_val(row_cells, col_map.get(key))
            if val is None or str(val).strip() == "" or str(val).lower() == "nan": return ""
            if key in ("dob", "doj", "doe"): return format_date(val)
            return str(val).strip()
            
        sl = None
        if "serial_no" in col_map:
            try: sl = int(float(get_val("serial_no")))
            except: pass
            
        records.append({
            "member_id": acc, 
            "name": name, 
            "wages": wages, 
            "gross_wages": gross_wages,
            "ncp_days": ncp_days,
            "uan": get_val("uan"),
            "father_name": get_val("father_name"),
            "dob": get_val("dob"),
            "sex": get_val("sex"),
            "doj": get_val("doj"),
            "doe": get_val("doe"),
            "reason_leaving": get_val("reason_leaving"),
            "serial_no": sl
        })

        def get_bool(key):
            v = cell_val(row_cells, col_map.get(key))
            if isinstance(v, bool): return v
            if isinstance(v, str) and v.lower().strip() in ('yes', 'y', 'true', '1'): return True
            if isinstance(v, (int, float)) and v == 1: return True
            return False

        records[-1]["gross_wages"] = gross_wages
        records[-1]["higher_epf"] = get_bool("higher_epf")
        records[-1]["age_crosses_58"] = get_bool("age_crosses_58")

    return records, warnings


def import_master_from_excel(filepath: str):
    """
    Reads Employee Master data using pandas. Supports .xlsx, .xls, and .csv.
    Understands EPFO standard headers (e.g. 'Member ID', 'UAN', 'Gender') as well as custom ones.
    """
    import pandas as pd
    import os
    
    ext = os.path.splitext(filepath)[1].lower()
    try:
        if ext == '.csv':
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
    except Exception as e:
        raise ValueError(f"Failed to read file using pandas: {e}")

    # Fallback if the first row isn't the header (common in some exports)
    # Check if 'Member ID' or 'Account' is in the columns
    header_found = False
    for col in df.columns:
        if isinstance(col, str) and ("account" in col.lower() or "member id" in col.lower()):
            header_found = True
            break
            
    if not header_found:
        # Scan all rows for header
        for idx in range(len(df)):
            row_vals = df.iloc[idx].astype(str).str.lower().tolist()
            has_id = any("account" in v or "member id" in v or "uan" in v for v in row_vals)
            has_name = any("name" in v for v in row_vals)
            if has_id and has_name:
                df.columns = df.iloc[idx]
                df = df.iloc[idx+1:].reset_index(drop=True)
                header_found = True
                break

    if not header_found:
        raise ValueError("Could not find a header row containing 'Member ID' or 'UAN'")

    col_map = {}
    for col in df.columns:
        text = str(col).lower().strip()
        if not text: continue
        if "member id" in text:
            col_map["member_id"] = col
        elif "account" in text and "bank" not in text and "member_id" not in col_map:
            col_map["member_id"] = col
        elif "uan" in text: col_map["uan"] = col
        elif "name" in text and "father" not in text and "husband" not in text: col_map["name"] = col
        elif "father" in text or "husband" in text: col_map["father_name"] = col
        elif "birth" in text or "dob" in text: col_map["dob"] = col
        elif "sex" in text or "gender" in text: col_map["sex"] = col
        elif "joining" in text or "doj" in text: col_map["doj"] = col
        elif "exit" in text or "leaving" in text and "reason" not in text: col_map["doe"] = col
        elif "reason" in text: col_map["reason_leaving"] = col
        elif text == "sl" or text == "sl no": col_map["serial_no"] = col
        elif "relationship" in text: col_map["relationship"] = col
        elif "marital" in text: col_map["marital_status"] = col
        elif "mobile" in text or "phone" in text: col_map["mobile"] = col
        elif "email" in text: col_map["email"] = col
        elif "aadhaar" in text: col_map["aadhaar"] = col
        elif "bank" in text and "account" in text: col_map["bank_account"] = col
        elif "ifsc" in text: col_map["ifsc"] = col
    def format_date(val):
        if pd.isna(val) or val == "": return ""
        if isinstance(val, str):
            val = val.strip().replace("/", "-")
            if val.endswith(" 00:00:00"): val = val.split(" ")[0]
            return val
        if hasattr(val, "strftime"):
            return val.strftime("%d-%m-%Y")
        return str(val)

    records = []
    warnings = []
    
    for _, row in df.iterrows():
        acc = row[col_map["member_id"]] if "member_id" in col_map else None
        if pd.isna(acc) or str(acc).strip() == "":
            uan_fallback = row[col_map["uan"]] if "uan" in col_map else None
            if not pd.isna(uan_fallback) and str(uan_fallback).strip() != "":
                acc = f"__UAN__{str(uan_fallback).strip()}"
            else:
                continue
        
        acc = str(acc).strip()
        name = str(row[col_map["name"]]).strip() if "name" in col_map and not pd.isna(row[col_map["name"]]) else ""
        
        def get_val(key):
            if key not in col_map: return ""
            v = row[col_map[key]]
            if pd.isna(v): return ""
            if key in ("dob", "doj", "doe"): return format_date(v)
            return str(v).strip()
            
        sl = None
        if "serial_no" in col_map:
            try: sl = int(float(get_val("serial_no")))
            except: pass
            
        records.append({
            "member_id": acc,
            "name": name,
            "uan": get_val("uan"),
            "father_name": get_val("father_name"),
            "dob": get_val("dob"),
            "sex": get_val("sex").capitalize(),
            "doj": get_val("doj"),
            "doe": get_val("doe"),
            "reason_leaving": get_val("reason_leaving"),
            "serial_no": sl,
            "relationship": get_val("relationship"),
            "marital_status": get_val("marital_status"),
            "mobile": get_val("mobile"),
            "email": get_val("email"),
            "aadhaar": get_val("aadhaar"),
            "bank_account": get_val("bank_account"),
            "ifsc": get_val("ifsc")
        })
        
    return records, warnings


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------

class ContributionSchemeMixin:
    """
    Shared rate logic for anything that carries a `scheme` plus the
    pre-1997 (epf_rate/fpf_rate) and post-1997 (emp_epf_rate/er_epf_rate/
    er_eps_rate) fields -- used by both Establishment and YearRecord so the
    math is defined in exactly one place.
    """

    @property
    def is_post_1997(self) -> bool:
        return self.scheme == SCHEME_POST_1997

    @property
    def worker_epf_rate(self) -> float:
        """Worker's EPF (Account 1) contribution rate, %."""
        return self.emp_epf_rate if self.is_post_1997 else self.epf_rate

    @property
    def worker_eps_rate(self) -> float:
        """Worker's Pension Fund contribution rate, %. Always 0 from
        1997-98 onwards -- the worker's whole share goes to EPF."""
        return 0.0 if self.is_post_1997 else self.fpf_rate

    @property
    def employer_epf_rate(self) -> float:
        """Employer's EPF (Account 1) contribution rate, %."""
        return self.er_epf_rate if self.is_post_1997 else self.epf_rate

    @property
    def employer_eps_rate(self) -> float:
        """Employer's Pension Fund/EPS (Account 10) contribution rate, %."""
        return self.er_eps_rate if self.is_post_1997 else self.fpf_rate

    @property
    def eps_label(self) -> str:
        return "PENSION FUND (EPS)" if self.is_post_1997 else "FPF"

    @property
    def statutory_rate(self) -> float:
        """Single headline % -- 12% or 10% entered in the Add Year function."""
        return round(self.emp_epf_rate, 2)

    @property
    def statutory_rate_text(self) -> str:
        """Human-readable statutory-rate line for the printed forms."""
        if self.is_post_1997:
            return (f"Employee: {self.emp_epf_rate:g}% (EPF)   |   "
                    f"Employer: {self.er_epf_rate:g}% (EPF) + {self.er_eps_rate:g}% (Pension Fund) "
                    f"= {self.er_epf_rate + self.er_eps_rate:g}%")
        return f"{self.statutory_rate:g}%"


@dataclass
class Establishment(ContributionSchemeMixin):
    code: str = ""              # e.g. "OR/1042576"
    name: str = ""               # e.g. "KANDARPUR COLLEGE"
    address: str = ""            # e.g. "AT/PO-KANDARPUR, CUTTACK"
    coverage_date: str = ""      # Date of Coverage under the EPF Act, DD/MM/YYYY -- shown on Form 9
    year_from: str = ""          # e.g. "1988"
    year_to: str = ""            # e.g. "1989"
    scheme: str = SCHEME_PRE_1997          # SCHEME_PRE_1997 or SCHEME_POST_1997
    epf_rate: float = 6.84       # PRE-1997 only: worker's/employer's EPF contribution %, of wages
    fpf_rate: float = 1.16       # PRE-1997 only: worker's/employer's FPF contribution %, of wages
    emp_epf_rate: float = 12.0   # POST-1997 only: worker's EPF contribution % (all goes to EPF)
    er_epf_rate: float = 3.67    # POST-1997 only: employer's EPF (Account 1) portion %
    er_eps_rate: float = 8.33    # POST-1997 only: employer's Pension Fund (Account 10) portion %

    @property
    def short_year_label(self) -> str:
        """e.g. '88-89' -- used for the 3A sheet name, matching EPFO convention."""
        if not (self.year_from and self.year_to):
            return ""
        return f"{self.year_from[-2:]}-{self.year_to[-2:]}"

    @property
    def long_year_label(self) -> str:
        """e.g. '1988-89' -- used for the 6A sheet name and headings."""
        if not (self.year_from and self.year_to):
            return ""
        return f"{self.year_from}-{self.year_to[-2:]}"

    def to_dict(self):
        return asdict(self)

    @staticmethod
    def from_dict(d):
        return Establishment(**d)


@dataclass
class Employee:
    member_id: str = ''
    name: str = ''
    father_name: str = ''
    uan: str = ''         # Universal Account Number -- shown on Form 3A alongside Member ID
    wages: List[float] = field(default_factory=lambda: [0.0] * 12)  # APR..MAR
    gross_wages: List[float] = field(default_factory=lambda: [0.0] * 12)
    ncp_days: List[int] = field(default_factory=lambda: [0] * 12)
    higher_epf_ee: bool = False
    higher_epf_er: bool = False
    age_crosses_58: bool = False

    def month_rows(self, worker_epf_rate: float, worker_eps_rate: float,
                   employer_epf_rate: float, employer_eps_rate: float,
                   wage_ceilings: List[float] = None):
        """
        Returns a list of 12 tuples:
            (wages, w_epf, w_eps, w_total, e_epf, e_eps, e_total)
        Rounded to the nearest rupee, exactly as the paper EPF forms do.
        """
        if wage_ceilings is None:
            wage_ceilings = [15000.0] * 12

        rows = []
        for i, w in enumerate(self.wages):
            w = w or 0
            ceiling = wage_ceilings[i]
            
            # Post-1997 calculation restrictions:
            if worker_eps_rate == 0:
                worker_wage_base = w if self.higher_epf_ee else min(w, ceiling)
                er_total_wage_base = w if self.higher_epf_er else min(w, ceiling)
                eps_wage = 0 if self.age_crosses_58 else min(w, ceiling)
                
                w_epf = round(worker_wage_base * worker_epf_rate / 100)
                w_eps = round(w * worker_eps_rate / 100)  # Will be 0 anyway
                
                e_eps = round(eps_wage * employer_eps_rate / 100)
                total_er_contrib = round(er_total_wage_base * worker_epf_rate / 100)
                e_epf = max(0, total_er_contrib - e_eps) 

            else:
                w_epf = round(w * worker_epf_rate / 100)
                w_eps = round(w * worker_eps_rate / 100)
                e_epf = round(w * employer_epf_rate / 100)
                e_eps = round(w * employer_eps_rate / 100)
                
            rows.append((w, w_epf, w_eps, w_epf + w_eps, e_epf, e_eps, e_epf + e_eps))
        return rows

    def annual_totals(self, worker_epf_rate: float, worker_eps_rate: float,
                       employer_epf_rate: float, employer_eps_rate: float,
                       wage_ceilings: List[float] = None):
        rows = self.month_rows(worker_epf_rate, worker_eps_rate, employer_epf_rate, employer_eps_rate, wage_ceilings)
        wages_total = sum(r[0] for r in rows)
        w_epf_total = sum(r[1] for r in rows)
        w_eps_total = sum(r[2] for r in rows)
        w_total_total = sum(r[3] for r in rows)
        e_epf_total = sum(r[4] for r in rows)
        e_eps_total = sum(r[5] for r in rows)
        e_total_total = sum(r[6] for r in rows)
        return (wages_total, w_epf_total, w_eps_total, w_total_total,
                e_epf_total, e_eps_total, e_total_total)

    def to_dict(self):
        return asdict(self)

    @staticmethod
    def from_dict(d):
        if "account_no" in d and "member_id" not in d:
            d["member_id"] = normalize_member_id(d.pop("account_no"))
        d["member_id"] = normalize_member_id(d.get("member_id", ""))
        return Employee(**d)


# --------------------------------------------------------------------------
# Employee Master -- member id / name / father's name, shared across years
# --------------------------------------------------------------------------

@dataclass
class MasterEmployee:
    member_id: str = ""
    name: str = ""
    father_name: str = ""
    uan: str = ""         # Universal Account Number
    dob: str = ""        # Date of Birth, DD/MM/YYYY
    sex: str = ""         # "Male" or "Female"
    doj: str = ""         # Date of Joining, DD/MM/YYYY
    doe: str = ""         # Date of Exit, DD/MM/YYYY
    reason_leaving: str = ""  # one of REASONS_FOR_LEAVING
    serial_no: int = 0    # SL No. -- the employee list sorts by THIS, not by member_id
    relationship: str = "" # Relationship to Father/Husband
    marital_status: str = ""
    mobile: str = ""
    email: str = ""
    aadhaar: str = ""
    bank_account: str = ""
    ifsc: str = ""
    higher_epf_ee: bool = False
    higher_epf_er: bool = False

    def to_dict(self):
        return asdict(self)

    @staticmethod
    def from_dict(d):
        mid = normalize_member_id(d.get("member_id") or d.get("account_no", ""))
        return MasterEmployee(
            member_id=mid, name=d.get("name", ""),
                               father_name=d.get("father_name", ""), uan=d.get("uan", ""),
                               dob=d.get("dob", ""), sex=d.get("sex", ""), doj=d.get("doj", ""),
                               doe=d.get("doe", ""), reason_leaving=d.get("reason_leaving", ""),
                               serial_no=d.get("serial_no", 0),
                               relationship=d.get("relationship", ""), marital_status=d.get("marital_status", ""),
                               mobile=d.get("mobile", ""), email=d.get("email", ""),
                               aadhaar=d.get("aadhaar", ""), bank_account=d.get("bank_account", ""),
                               ifsc=d.get("ifsc", ""),
                               higher_epf_ee=d.get("higher_epf_ee", False),
                               higher_epf_er=d.get("higher_epf_er", False))

    @property
    def age_years(self):
        return calc_age_years(self.dob)

    @property
    def is_due_superannuation(self):
        age = self.age_years
        return age is not None and age >= SUPERANNUATION_AGE


@dataclass
class YearEntry:
    """One employee's wage entry for one specific year."""
    member_id: str = ""
    wages: List[float] = field(default_factory=lambda: [0.0] * 12)  # APR..MAR
    gross_wages: List[float] = field(default_factory=lambda: [0.0] * 12)
    ncp_days: List[int] = field(default_factory=lambda: [0] * 12)
    age_crosses_58: bool = False

    def to_dict(self):
        return asdict(self)

    @staticmethod
    def from_dict(d):
        d.pop("higher_epf", None)  # Safe cleanup
        if "account_no" in d and "member_id" not in d:
            d["member_id"] = normalize_member_id(d.pop("account_no"))
        d["member_id"] = normalize_member_id(d.get("member_id", ""))
        return YearEntry(**d)


@dataclass
class YearRecord(ContributionSchemeMixin):
    year_from: str = ""
    year_to: str = ""
    scheme: str = SCHEME_PRE_1997          # SCHEME_PRE_1997 or SCHEME_POST_1997
    epf_rate: float = 6.84       # PRE-1997 only
    fpf_rate: float = 1.16       # PRE-1997 only
    emp_epf_rate: float = 12.0   # POST-1997 only: worker's EPF %
    er_epf_rate: float = 3.67    # POST-1997 only: employer's EPF portion %
    er_eps_rate: float = 8.33    # POST-1997 only: employer's Pension Fund portion %
    entries: List[YearEntry] = field(default_factory=list)
    remittances: List[dict] = field(default_factory=list)

    @property
    def long_label(self) -> str:
        if not (self.year_from and self.year_to):
            return ""
        return f"{self.year_from}-{self.year_to[-2:]}"

    @property
    def short_label(self) -> str:
        if not (self.year_from and self.year_to):
            return ""
        return f"{self.year_from[-2:]}-{self.year_to[-2:]}"

    def to_dict(self):
        d = asdict(self)
        return d

    @staticmethod
    def from_dict(d):
        entries = [YearEntry.from_dict(e) for e in d.get("entries", [])]
        return YearRecord(year_from=d.get("year_from", ""), year_to=d.get("year_to", ""),
                           scheme=d.get("scheme", SCHEME_PRE_1997),
                           epf_rate=d.get("epf_rate", 6.84), fpf_rate=d.get("fpf_rate", 1.16),
                           emp_epf_rate=d.get("emp_epf_rate", 12.0),
                           er_epf_rate=d.get("er_epf_rate", 3.67),
                           er_eps_rate=d.get("er_eps_rate", 8.33),
                           entries=entries,
                           remittances=d.get("remittances", []))


# --------------------------------------------------------------------------
# Project: one Establishment + a shared Employee Master + many years
# --------------------------------------------------------------------------

class Project:
    def __init__(self):
        self.code = ""
        self.name = ""
        self.address = ""
        self.coverage_date = ""         # Date of Coverage under the EPF Act, DD/MM/YYYY
        self.created_at = datetime.now().strftime("%d-%m-%Y")
        self.is_active = True
        self.master: dict = {}          # member_id -> MasterEmployee
        self.years: dict = {}           # year_key (long_label) -> YearRecord
        self.current_year_key = None

    # ---- establishment ----
    def set_establishment(self, code, name, address, coverage_date=""):
        self.code, self.name, self.address, self.coverage_date = code, name, address, coverage_date

    # ---- employee master ----
    def resolve_member_id(self, member_id: str, uan: str) -> str:
        if not uan:
            return normalize_member_id(member_id)
        
        uan = str(uan).strip()
        for m_id, emp in self.master.items():
            if str(emp.uan).strip() == uan:
                return m_id
        
        return normalize_member_id(member_id)

    def upsert_master(self, member_id, name, father_name="", uan="", dob="", sex="", doj="",
                       doe="", reason_leaving="", serial_no=None, relationship="", marital_status="",
                       mobile="", email="", aadhaar="", bank_account="", ifsc="",
                       higher_epf_ee=False, higher_epf_er=False):
        member_id = normalize_member_id(member_id)
        if member_id in self.master:
            m = self.master[member_id]
            m.name = name or m.name
            if father_name: m.father_name = father_name
            if uan: m.uan = uan
            if dob: m.dob = dob
            if sex: m.sex = sex
            if doj: m.doj = doj
            if doe: m.doe = doe
            if reason_leaving: m.reason_leaving = reason_leaving
            if serial_no is not None: m.serial_no = serial_no
            if relationship: m.relationship = relationship
            if marital_status: m.marital_status = marital_status
            if mobile: m.mobile = mobile
            if email: m.email = email
            if aadhaar: m.aadhaar = aadhaar
            if bank_account: m.bank_account = bank_account
            if ifsc: m.ifsc = ifsc
            m.higher_epf_ee = higher_epf_ee
            m.higher_epf_er = higher_epf_er
        else:
            if serial_no is None:
                serial_no = self.next_serial_no()
            self.master[member_id] = MasterEmployee(member_id=member_id, name=name, father_name=father_name,
                                                       uan=uan, dob=dob, sex=sex, doj=doj, doe=doe,
                                                       reason_leaving=reason_leaving, serial_no=serial_no,
                                                       relationship=relationship, marital_status=marital_status,
                                                       mobile=mobile, email=email, aadhaar=aadhaar,
                                                       bank_account=bank_account, ifsc=ifsc,
                                                       higher_epf_ee=higher_epf_ee, higher_epf_er=higher_epf_er)

    def next_serial_no(self):
        """Next SL No. suggestion for a brand-new employee (one more than the
        highest SL No. currently in use)."""
        existing = [m.serial_no for m in self.master.values() if m.serial_no]
        return (max(existing) + 1) if existing else 1

    def get_master(self, member_id):
        member_id = normalize_member_id(member_id)
        return self.master.get(member_id.strip())

    def remove_master(self, member_id):
        member_id = normalize_member_id(member_id)
        self.master.pop(member_id, None)

    def rename_account(self, old_acc, new_acc):
        """Rename an employee's account number across the Employee Master and
        all year wage entries, so wage data is preserved when an account number
        is corrected or replaced (e.g. with a UAN)."""
        old_acc = old_acc.strip()
        new_acc = new_acc.strip()
        if old_acc in self.master:
            m = self.master.pop(old_acc)
            m.member_id = new_acc
            self.master[new_acc] = m
        for yr in self.years.values():
            for e in yr.entries:
                if e.member_id == old_acc:
                    e.member_id = new_acc

    def master_list(self):
        """Employee Master list, sorted by SL No. (serial_no) -- NOT by
        account number. Employees without a serial_no (e.g. very old
        projects saved before this field existed) sort to the end, in
        natural account-number order among themselves."""
        return sorted(self.master.values(),
                      key=lambda m: (m.serial_no if m.serial_no else float("inf"),
                                     natural_sort_key(m.member_id)))

    # ---- years ----
    def add_year(self, year_from, year_to, scheme=SCHEME_PRE_1997,
                 epf_rate=6.84, fpf_rate=1.16,
                 emp_epf_rate=12.0, er_epf_rate=3.67, er_eps_rate=8.33):
        yr = YearRecord(year_from=year_from, year_to=year_to, scheme=scheme,
                         epf_rate=epf_rate, fpf_rate=fpf_rate,
                         emp_epf_rate=emp_epf_rate, er_epf_rate=er_epf_rate, er_eps_rate=er_eps_rate)
        key = yr.long_label
        self.years[key] = yr
        self.current_year_key = key
        return key

    def update_year_rates(self, key, scheme, epf_rate, fpf_rate,
                          emp_epf_rate, er_epf_rate, er_eps_rate):
        if key in self.years:
            yr = self.years[key]
            yr.scheme = scheme
            yr.epf_rate = epf_rate
            yr.fpf_rate = fpf_rate
            yr.emp_epf_rate = emp_epf_rate
            yr.er_epf_rate = er_epf_rate
            yr.er_eps_rate = er_eps_rate

    def remove_year(self, key):
        self.years.pop(key, None)
        if self.current_year_key == key:
            self.current_year_key = next(iter(self.years), None)

    def year_keys_sorted(self):
        return sorted(self.years.keys(), key=lambda k: self.years[k].year_from)

    def current_year(self) -> YearRecord:
        if self.current_year_key and self.current_year_key in self.years:
            return self.years[self.current_year_key]
        return None

    # ---- entries (employee wages for a given year) ----
    def get_entry(self, year_key, member_id):
        yr = self.years.get(year_key)
        if not yr: return None
        member_id = normalize_member_id(member_id)
        for e in yr.entries:
            if e.member_id == member_id:
                return e
        return None


    def upsert_entry(self, year_key, member_id, wages, gross_wages=None, ncp_days=None, age_crosses_58=False, higher_epf_ee=None, higher_epf_er=None):
        yr = self.years[year_key]
        member_id = normalize_member_id(member_id)
        if gross_wages is None: gross_wages = wages.copy()
        if ncp_days is None: ncp_days = [0] * 12
        
        m = self.get_master(member_id)
        if m:
            if higher_epf_ee is not None:
                m.higher_epf_ee = higher_epf_ee
            if higher_epf_er is not None:
                m.higher_epf_er = higher_epf_er
        for e in yr.entries:
            if e.member_id == member_id:
                e.wages = wages
                e.gross_wages = gross_wages
                e.ncp_days = ncp_days
                e.age_crosses_58 = age_crosses_58
                return
        yr.entries.append(YearEntry(member_id=member_id, wages=wages, gross_wages=gross_wages, ncp_days=ncp_days, age_crosses_58=age_crosses_58))

    def remove_entry(self, year_key, index):
        del self.years[year_key].entries[index]

    def build_employees_for_year(self, year_key) -> List[Employee]:
        """Merges each year-entry's wages with the employee's master name/father,
        producing the flat Employee list the ExcelGenerator understands.
        Ordered by the Employee Master's SL No. (serial_no) -- the same
        order the Employee Master list is shown in -- not by account number.
        Entries whose account number isn't in the Master (shouldn't normally
        happen) sort to the end, in natural account-number order."""
        yr = self.years[year_key]

        def sort_key(e):
            m = self.master.get(e.member_id)
            serial = m.serial_no if (m and m.serial_no) else float("inf")
            return (serial, natural_sort_key(e.member_id))

        result = []
        for e in sorted(yr.entries, key=sort_key):
            m = self.master.get(e.member_id)
            name = m.name if m else ""
            father = m.father_name if m else ""
            uan = m.uan if m else ""
            result.append(Employee(member_id=e.member_id, name=name, father_name=father, uan=uan,
                                    wages=list(e.wages), gross_wages=list(e.gross_wages), ncp_days=list(getattr(e, 'ncp_days', [0]*12)),
                                    higher_epf_ee=m.higher_epf_ee if m else False,
                                    higher_epf_er=m.higher_epf_er if m else False,
                                    age_crosses_58=e.age_crosses_58))
        return result

    def build_establishment_for_year(self, year_key) -> Establishment:
        yr = self.years[year_key]
        return Establishment(code=self.code, name=self.name, address=self.address,
                              coverage_date=self.coverage_date,
                              year_from=yr.year_from, year_to=yr.year_to,
                              scheme=yr.scheme, epf_rate=yr.epf_rate, fpf_rate=yr.fpf_rate,
                              emp_epf_rate=yr.emp_epf_rate, er_epf_rate=yr.er_epf_rate,
                              er_eps_rate=yr.er_eps_rate)

    # ---- persistence ----
    def to_dict(self):
        return {
            "code": self.code, "name": self.name, "address": self.address,
            "coverage_date": self.coverage_date,
            "created_at": getattr(self, "created_at", datetime.now().strftime("%d-%m-%Y")),
            "is_active": getattr(self, "is_active", True),
            "master": {k: v.to_dict() for k, v in self.master.items()},
            "years": {k: v.to_dict() for k, v in self.years.items()},
            "current_year_key": self.current_year_key,
        }

    def load_from_dict(self, data: dict):
        self.code = data.get("code", "")
        self.name = data.get("name", "")
        self.address = data.get("address", "")
        self.coverage_date = data.get("coverage_date", "")
        self.created_at = data.get("created_at", datetime.now().strftime("%d-%m-%Y"))
        self.is_active = data.get("is_active", True)
        self.master = {normalize_member_id(k): MasterEmployee.from_dict(v) for k, v in data.get("master", {}).items()}
        self.years = {normalize_member_id(k): YearRecord.from_dict(v) for k, v in data.get("years", {}).items()}
        self.current_year_key = data.get("current_year_key") or next(iter(self.years), None)
        
        # Migrate old 'higher_epf' from entries to MasterEmployee
        for yr_data in data.get("years", {}).values():
            for e_data in yr_data.get("entries", []):
                if e_data.get("higher_epf"):
                    mid = normalize_member_id(e_data.get("account_no") or e_data.get("member_id", ""))
                    if mid in self.master:
                        self.master[mid].higher_epf_ee = True
                        self.master[mid].higher_epf_er = True

    def save(self, filepath: str):
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)

    def load(self, filepath: str):
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.load_from_dict(data)

    def new(self):
        self.__init__()

    # ---- plain-data Excel export (for viewing/copying -- NOT the official
    #      Form 3A/6A, which ExcelGenerator produces separately) ----
    def export_data_workbook(self, filepath: str):
        """
        Writes a single, plain Excel workbook that mirrors everything stored
        in this project:
          - "Employee Master" sheet: every employee and all their master details.
          - One "Wages_<year>" sheet per year: every employee's 12 monthly
            wages plus the computed EPF/EPS contributions for that year.

        This is a flat reference copy meant to be opened, sorted, filtered,
        and copied from directly in Excel for official record-keeping -- it
        is separate from the printable Form 3A / Form 6A output.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9D9D9")

        # ---- Employee Master sheet ----
        ws = wb.create_sheet("Employee Master")
        headers = ["SL", "Member ID", "UAN No.", "Name", "Father's Name", "Date of Birth", "Age",
                   "Sex", "Date of Joining", "Date of Exit", "Reason of Leaving"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
        for i, m in enumerate(self.master_list(), start=1):
            age = calc_age_years(m.dob)
            ws.append([i, m.member_id, m.uan, m.name, m.father_name, m.dob,
                       age if age is not None else "", m.sex, m.doj, m.doe, m.reason_leaving])
        for col, width in zip("ABCDEFGHIJK", [5, 16, 16, 24, 22, 14, 6, 8, 14, 14, 22]):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"

        # ---- one sheet per year, oldest first ----
        for year_key in sorted(self.years.keys(),
                                key=lambda k: natural_sort_key(self.years[k].year_from)):
            yr = self.years[year_key]
            est = self.build_establishment_for_year(year_key)
            employees = self.build_employees_for_year(year_key)

            sheet_name = f"Wages_{yr.short_label}"[:31]
            ws = wb.create_sheet(sheet_name)
            scheme_label = "Post-1997" if yr.is_post_1997 else "Pre-1997"
            ws.append([f"Year: {yr.long_label}", f"Scheme: {scheme_label}", est.statutory_rate_text])
            ws["A1"].font = Font(bold=True, italic=True)

            headers2 = (["SL", "Member ID", "UAN No.", "Name"] + MONTHS + ["Annual Wages",
                        "Worker EPF", f"Worker {est.eps_label}", "Worker Total",
                        "Employer EPF", f"Employer {est.eps_label}", "Employer Total"])
            ws.append(headers2)
            for c in range(1, len(headers2) + 1):
                cell = ws.cell(row=2, column=c)
                cell.font = header_font
                cell.fill = header_fill

            for i, emp in enumerate(employees, start=1):
                wt, w_epf, w_eps, w_tot, e_epf, e_eps, e_tot = emp.annual_totals(
                    est.worker_epf_rate, est.worker_eps_rate, est.employer_epf_rate, est.employer_eps_rate)
                row = ([i, emp.member_id, emp.uan, emp.name] + [w or 0 for w in emp.wages] +
                       [wt, w_epf, w_eps, w_tot, e_epf, e_eps, e_tot])
                ws.append(row)

            ws.freeze_panes = "D3"
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 22

        wb.save(filepath)


# --------------------------------------------------------------------------
# Excel generation (Form 3A + Form 6A), styled to match the official EPFO forms
# --------------------------------------------------------------------------

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", bold=True, size=10, color="1F4E78")
BOLD_ITALIC = Font(name="Arial", bold=True, italic=True, size=10)
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _wrapped_line_count(text, col_width_chars):
    """
    Estimates how many display lines `text` will wrap to inside a column
    (or merged span) that's col_width_chars wide (Excel's column-width
    unit is roughly one character). Explicit "\n" line breaks are honoured
    as forced line breaks; each resulting segment is further wrapped if
    it's still too long to fit in one line.
    """
    if not text:
        return 1
    total = 0
    for seg in str(text).split("\n"):
        seg = seg.strip()
        total += max(1, -(-len(seg) // max(1, int(col_width_chars))))  # ceil division
    return max(1, total)


def _row_height_for_cells(cells, line_height=13, padding=6, min_height=15):
    """
    cells: list of (text, col_width_chars) pairs -- one per header cell in a
    row (use the summed width of every column a merged cell spans). Returns
    a row height (points) tall enough that the cell needing the most lines
    isn't clipped, so table headers stay fully readable instead of wrapping
    into invisible overflow. Used for every table header in Form 3A, 6A,
    12A, Form 5 and Form 10, plus the "Certified that..." paragraphs on
    Form 3A.
    """
    max_lines = max((_wrapped_line_count(text, width) for text, width in cells), default=1)
    return max(min_height, max_lines * line_height + padding)


# --------------------------------------------------------------------------
# Form 12A: statutory admin-charge rates, which have changed several times
# over the years. Applied per the EXACT calendar month (not per financial
# year), since a single financial year can straddle a rate change --
# e.g. FY 2018-19 (Apr 2018-Mar 2019) changed mid-year in June 2018.
# --------------------------------------------------------------------------
ACCOUNT_21_RATE = 0.50   # % of monthly total wages -- EDLI contribution (A/c 21). Unchanged historically.
ACCOUNT_22_MIN = 2       # minimum Rs. 2/month for A/c 22, when a charge applies

_MONTH_NUM = {"APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9,
              "OCT": 10, "NOV": 11, "DEC": 12, "JAN": 1, "FEB": 2, "MAR": 3}

def get_month_num(label: str) -> int:
    return _MONTH_NUM.get(label[:3].upper(), 0)

def calendar_year_for_month(month_abbr: str, year_from: str, year_to: str):
    """MONTHS runs Mar..Feb; Mar-Dec belong to year_from, Jan-Feb to year_to."""
    m = _MONTH_NUM.get(month_abbr[:3].upper())
    if not m:
        return None
    try:
        yf = int(str(year_from)[:4]) if year_from else None
        yt = int(str(year_to)[:4]) if year_to else None
    except ValueError:
        return None
    return yf if m >= 3 else yt


def account2_rate_percent(cal_year, cal_month):
    """
    EPF Administrative/Inspection charges (A/c No. 2), % of that month's
    total EPF wages:
      Pre-2015:            1.10%
      Jan 2015 - Mar 2017:  0.85%
      Apr 2017 - May 2018:  0.65%
      Jun 2018 onwards:     0.50%
    """
    if cal_year is None:
        return 1.10
    ym = (cal_year, cal_month)
    if ym < (2015, 1):
        return 1.10
    elif ym < (2017, 4):
        return 0.85
    elif ym < (2018, 6):
        return 0.65
    else:
        return 0.50


def account22_rate_percent(cal_year, cal_month):
    """EDLI Administrative charges (A/c No. 22) -- waived (0%) from 1 April
    2017 onwards; 0.01% (subject to a Rs. 2/month minimum) before that."""
    if cal_year is None:
        return 0.01
    return 0.01 if (cal_year, cal_month) < (2017, 4) else 0.0


def format_rate_periods(month_rate_pairs):
    """
    month_rate_pairs: [(month_abbr, rate), ...] in calendar (Apr..Mar) order.
    Groups consecutive months sharing the same rate into readable segments,
    e.g. "APR-MAY: 0.65%, JUN-MAR: 0.5%" -- used in the Form 12A footnote so
    it's clear which rate applied when a financial year straddles a change.
    """
    if not month_rate_pairs:
        return ""
    segments = []
    start_month, start_rate = month_rate_pairs[0]
    prev_month, prev_rate = month_rate_pairs[0]
    for month, rate in month_rate_pairs[1:]:
        if rate != prev_rate:
            label = start_month if start_month == prev_month else f"{start_month}-{prev_month}"
            segments.append(f"{label}: {prev_rate:g}%")
            start_month, start_rate = month, rate
        prev_month, prev_rate = month, rate
    label = start_month if start_month == prev_month else f"{start_month}-{prev_month}"
    segments.append(f"{label}: {prev_rate:g}%")
    return ", ".join(segments)



class ExcelGenerator:
    def __init__(self, establishment: Establishment, employees: List[Employee], project: "Project" = None, forms_to_generate: List[str] = None):
        """
        project is optional and only needed to also emit the Form 5 / Form 10
        monthly sheets (they're built from the Employee Master's Date of
        Joining / Date of Exit, which live on the project, not on a single
        year's Employee list). If project is omitted, only 3A/6A/12A are built
        -- exactly the previous behaviour.
        """
        self.est = establishment
        self.employees = [emp for emp in employees if sum(emp.wages) > 0]
        self.project = project
        self.forms_to_generate = forms_to_generate or ['3A', '6A', '12A', '5', '10']

    def build(self, filepath: str):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        if '3A' in self.forms_to_generate:
            self._build_3a_sheet(wb)
        if '6A' in self.forms_to_generate:
            self._build_6a_sheet(wb)
        if '12A' in self.forms_to_generate:
            self._build_12a_sheet(wb)
        if self.project is not None:
            if '5' in self.forms_to_generate or '10' in self.forms_to_generate:
                self._build_form5_form10_sheets(wb, forms_to_generate=self.forms_to_generate)
        if not wb.sheetnames:
            wb.create_sheet("Empty")
        wb.save(filepath)
        return filepath

    @staticmethod
    def _write_signature_block(ws, row, num_cols, left_col=1):
        """
        Writes the standard "Signature of the Employer with seal" line
        (right-aligned) and "Date" on the left side of the SAME line.
        """
        ws.cell(row=row, column=left_col, value="Date").font = NORMAL
        
        if num_cols > left_col:
            ws.merge_cells(start_row=row, start_column=left_col + 1, end_row=row, end_column=num_cols)
            c = ws.cell(row=row, column=left_col + 1, value="Signature of the Employer with seal")
            c.font = NORMAL
            c.alignment = Alignment(horizontal="right")
            
        row += 1
        return row

    @staticmethod
    def _apply_a4_page_setup(ws, last_row, num_cols=11, orientation="landscape", margins=None,
                              fit_one_page=False, center_on_page=False):
        """
        Makes a sheet print cleanly on A4: landscape (the tables are wide),
        scaled to fit the page width (so nothing gets cut off sideways),
        with a sensible print area and slim margins. Used for both Form 3A
        and Form 6A, and honoured automatically if the workbook is later
        converted to PDF (see export_workbook_to_pdf).

        fit_one_page=True also scales to fit the page HEIGHT, forcing the
        whole sheet onto a single printed page instead of spilling a
        near-empty extra page (used for Form 5 / Form 10 / Form 12A, which
        are always short enough to fit).

        center_on_page=True centers the printed content horizontally on the
        page, instead of it hugging the left margin -- used for Form 5,
        Form 10 and Form 12A.
        """
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = orientation
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1 if fit_one_page else 0  # 0 = flow to as many pages tall as needed
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = margins or PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                                  header=0.2, footer=0.2)
        last_col = get_column_letter(num_cols)
        ws.print_area = f"A1:{last_col}{last_row}"
        if center_on_page:
            ws.print_options.horizontalCentered = True

    # ---------------------------------------------------------------- 3A ---
    def _build_3a_sheet(self, wb):
        est = self.est
        sheet_name = f"3A_{est.short_year_label}" if est.short_year_label else "3A"
        ws = wb.create_sheet(title=sheet_name[:31])
        self._3a_sheet_name = sheet_name[:31]
        self._3a_total_rows = []  # total row number for each employee, in order

        col_widths = [16, 14, 12, 10, 12, 10, 10, 12, 10, 12, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 2  # first block starts a little below the top, like the source file
        total_employees = len(self.employees)
        for idx, emp in enumerate(self.employees):
            row, total_row = self._write_3a_block(ws, row, emp)
            self._3a_total_rows.append(total_row)
            if idx < total_employees - 1:
                # Force a page break here so the NEXT employee's card starts
                # fresh on its own page, instead of Excel/PDF splitting a
                # card's 12-month table across two pages wherever the page
                # height happens to run out.
                ws.row_breaks.append(Break(id=row - 1))

        # Tighter margins than the default helper (0.3in vs 0.4/0.5in) buy a
        # bit more usable height, on top of the compacted row layout above,
        # so a full card (header + 12 months + certification + signature)
        # reliably fits within one A4 landscape page.
        tight_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3, header=0.15, footer=0.15)
        self._apply_a4_page_setup(ws, last_row=max(row - 1, 1), num_cols=len(col_widths),
                                   margins=tight_margins)

    def _write_3a_block(self, ws, start_row, emp: Employee):
        """Writes one employee's Form 3A card starting at start_row. Returns the
        row *after* the block (i.e. the next free row)."""
        est = self.est
        r = start_row

        def merged(row, text, font=NORMAL, align=CENTER, end_col=11):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
            c = ws.cell(row=row, column=1, value=text)
            c.font = font
            c.alignment = align
            return c

        merged(r, "(For Un-exempted Establishments only)", NORMAL, CENTER); r += 1
        merged(r, "FORM - 3A(R)", TITLE_FONT, CENTER); r += 1
        merged(r, "THE EMPLOYEE'S PROVIDENT FUND SCHEME, 1952 (PARA 35 & 43)", SUBTITLE_FONT, CENTER); r += 1
        merged(r, "THE EMPLOYEE'S PENSION SCHEME, 1995 (PARAGRAPH 20(4))", SUBTITLE_FONT, CENTER); r += 1
        merged(r, f"Contribution Card for the currency period April, {est.year_from} to March, {est.year_to}",
               BOLD_ITALIC, CENTER); r += 1

        member_id_row = r
        ws.cell(row=r, column=1, value=1).font = NORMAL
        ws.cell(row=r, column=2, value="Member ID").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        member_id_display_val = format_member_id_display(emp.member_id)
        member_id_display = f"{member_id_display_val}  (UAN: {emp.uan})" if emp.uan else member_id_display_val
        ws.cell(row=r, column=6, value=member_id_display).font = BOLD
        r += 1

        ws.cell(row=r, column=1, value=2).font = NORMAL
        ws.cell(row=r, column=2, value="Name of the Member:").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        ws.cell(row=r, column=6, value=emp.name).font = BOLD
        r += 1

        ws.cell(row=r, column=1, value=3).font = NORMAL
        ws.cell(row=r, column=2, value="Father's Name").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        ws.cell(row=r, column=6, value=emp.father_name).font = NORMAL
        r += 1

        ws.cell(row=r, column=1, value=4).font = NORMAL
        ws.cell(row=r, column=2, value="Name & Address of the Establishment").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        ws.cell(row=r, column=6, value=f"{est.name}, {est.address}").font = NORMAL
        r += 1
        ws.cell(row=r, column=2, value="Code No. of the Establishment").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        ws.cell(row=r, column=6, value=est.code).font = BOLD
        r += 1

        ws.cell(row=r, column=1, value=5).font = NORMAL
        ws.cell(row=r, column=2, value="Statutory Rate of Contribution").font = NORMAL
        ws.cell(row=r, column=5, value=":").font = NORMAL
        if est.is_post_1997:
            ws.cell(row=r, column=6, value=est.statutory_rate_text).font = NORMAL
        else:
            ws.cell(row=r, column=6, value=est.statutory_rate / 100).number_format = "0%"
        r += 1
        ws.cell(row=r, column=2, value="Voluntary higher rate of employee's contribution, if any").font = NORMAL
        r += 1

        # table header
        header_row1 = r
        header_row2 = r + 1

        ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
        ws.cell(row=header_row1, column=1, value="Month")
        
        ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row2, end_column=2)
        ws.cell(row=header_row1, column=2, value="Wages")

        ws.merge_cells(start_row=header_row1, start_column=3, end_row=header_row1, end_column=5)
        ws.cell(row=header_row1, column=3, value="WORKER'S SHARE")
        
        ws.merge_cells(start_row=header_row1, start_column=6, end_row=header_row1, end_column=8)
        ws.cell(row=header_row1, column=6, value="EMPLOYER'S SHARE")
        
        ws.merge_cells(start_row=header_row1, start_column=9, end_row=header_row2, end_column=9)
        ws.cell(row=header_row1, column=9, value="REFUND OF\nADVANCES")
        
        ws.merge_cells(start_row=header_row1, start_column=10, end_row=header_row2, end_column=10)
        ws.cell(row=header_row1, column=10, value="NCP\nDAYS")
        
        ws.merge_cells(start_row=header_row1, start_column=11, end_row=header_row2, end_column=11)
        ws.cell(row=header_row1, column=11, value="REMARKS")
        r += 1

        w_epf_rate, w_eps_rate = est.worker_epf_rate, est.worker_eps_rate
        e_epf_rate, e_eps_rate = est.employer_epf_rate, est.employer_eps_rate
        eps_label = est.eps_label
        
        headers2_inner = [f"EPF {w_epf_rate:g}%", f"{eps_label} {w_eps_rate:g}%", "TOTAL",
                          f"EPF {e_epf_rate:g}%", f"{eps_label} {e_eps_rate:g}%", "TOTAL"]
        for i, h in enumerate(headers2_inner, start=3):
            ws.cell(row=header_row2, column=i, value=h)
        r += 1

        for row_i in (header_row1, header_row2):
            for col in range(1, 12):
                c = ws.cell(row=row_i, column=col)
                c.font = BOLD
                c.fill = HEADER_FILL
                c.border = BORDER
                c.alignment = CENTER

        col_widths_3a = [16, 14, 12, 10, 12, 10, 10, 12, 10, 12, 12]
        ws.row_dimensions[header_row1].height = _row_height_for_cells([
            ("Month", col_widths_3a[0]),
            ("Wages", col_widths_3a[1]),
            ("WORKER'S SHARE", sum(col_widths_3a[2:5])),
            ("EMPLOYER'S SHARE", sum(col_widths_3a[5:8])),
            ("REFUND OF\nADVANCES", col_widths_3a[8]),
            ("NCP\nDAYS", col_widths_3a[9]),
            ("REMARKS", col_widths_3a[10]),
        ])
        
        headers2_full = ["", ""] + headers2_inner + ["", "", ""]
        ws.row_dimensions[header_row2].height = _row_height_for_cells(list(zip(headers2_full, col_widths_3a)))

        # month rows -- computed directly in Python (not live Excel formulas),
        # so the figures always display correctly regardless of the viewer's
        # recalculation settings.
        month_rows = emp.month_rows(w_epf_rate, w_eps_rate, e_epf_rate, e_eps_rate)
        first_month_row = r
        for i, m in enumerate(MONTHS):
            wages, w_epf, w_eps, w_total, e_epf, e_eps, e_total = month_rows[i]
            ws.cell(row=r, column=1, value=m).font = NORMAL
            ws.cell(row=r, column=2, value=wages).font = NORMAL
            ws.cell(row=r, column=3, value=w_epf).font = NORMAL
            ws.cell(row=r, column=4, value=w_eps).font = NORMAL
            ws.cell(row=r, column=5, value=w_total).font = NORMAL
            ws.cell(row=r, column=6, value=e_epf).font = NORMAL
            ws.cell(row=r, column=7, value=e_eps).font = NORMAL
            ws.cell(row=r, column=8, value=e_total).font = NORMAL
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = BORDER
                if col >= 2:
                    ws.cell(row=r, column=col).alignment = RIGHT
            r += 1

        last_month_row = r - 1
        total_row = r
        wt, w_epf_t, w_eps_t, w_tot_t, e_epf_t, e_eps_t, e_tot_t = emp.annual_totals(
            w_epf_rate, w_eps_rate, e_epf_rate, e_eps_rate)
        ws.cell(row=total_row, column=1, value="Total").font = BOLD
        for col_idx, val in zip(range(2, 9), [wt, w_epf_t, w_eps_t, w_tot_t, e_epf_t, e_eps_t, e_tot_t]):
            c = ws.cell(row=total_row, column=col_idx, value=val)
            c.font = BOLD
        for col in range(1, 12):
            ws.cell(row=total_row, column=col).border = BORDER
            if col >= 2:
                ws.cell(row=total_row, column=col).alignment = RIGHT
        r += 1

        cert1 = ("Certified that the total amount of contribution indicated in this card has already "
                 "been remitted in full in EPF A/c. No. 1 and A/c No. 10 vide note below.")
        merged(r, cert1, NOTE_FONT, LEFT)
        ws.row_dimensions[r].height = _row_height_for_cells([(cert1, 130)], line_height=11, min_height=13)
        r += 1
        merged(r, "(a) Date of leaving Service:                                (b) Reason for leaving service:", NORMAL, LEFT)
        r += 1
        cert2 = ("Certified that the difference between the total of the contributions shown under Cols. 3 & 4 "
                 "of the above table and that arrived at on the total wages shown in Col. 2 at the prescribed "
                 "rate is solely due to the rounding off of contribution to the nearest rupee under the rules.")
        merged(r, cert2, NOTE_FONT, LEFT)
        ws.row_dimensions[r].height = _row_height_for_cells([(cert2, 130)], line_height=11, min_height=13)
        r += 1
        r = ExcelGenerator._write_signature_block(ws, r, num_cols=11, left_col=1)

        return r, total_row

    # ---------------------------------------------------------------- 6A ---
    def _build_6a_sheet(self, wb):
        est = self.est
        sheet_name = f"6A_{est.short_year_label}" if est.short_year_label else "6A"
        ws = wb.create_sheet(title=sheet_name[:31])

        ws.merge_cells("A1:K1")
        ws["A1"] = "FORM 6 A"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:K2")
        ws["A2"] = "THE EMPLOYEE'S PROVIDENT FUND, 1952 (PARAGRAPH 43)"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:K3")
        ws["A3"] = "THE EMPLOYEE'S PENSION SCHEME, 1995 (PARAGRAPH 20(4))"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:K4")
        ws["A4"] = (f"Annual Statement of Contribution for the currency period from "
                    f"1st April {est.year_from} to 31st March {est.year_to}")
        ws["A4"].font = BOLD_ITALIC
        ws["A4"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A5:K5")
        ws["A5"] = f"Name & Address of the Establishment :- {est.name}, {est.address}"
        ws["A5"].font = Font(name="Arial", bold=True, italic=True, size=11)
        ws["A5"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A6:K6")
        ws["A6"] = f"Code No. of the Establishment :- {est.code}"
        ws["A6"].font = Font(name="Arial", bold=True, size=11)
        ws["A6"].alignment = Alignment(horizontal="center")

        ws.merge_cells("F7:J7")
        ws["F7"] = f"Statutory Rate of Contribution : {est.statutory_rate_text}"
        ws["F7"].font = BOLD_ITALIC
        ws["F7"].alignment = Alignment(horizontal="center")

        header_row1, header_row2, num_row = 8, 9, 10

        def span(r1, c1, r2, c2, text):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            ws.cell(row=r1, column=c1, value=text)

        span(header_row1, 1, header_row2, 1, "SL.NO")
        span(header_row1, 2, header_row2, 2, "MEMBER ID")
        span(header_row1, 3, header_row2, 3, "NAME OF EMPLOYEES")
        span(header_row1, 4, header_row2, 4, "WAGES BASIC Rs.")
        span(header_row1, 5, header_row1, 6, "WORKER'S CONTRIBUTION")
        span(header_row1, 7, header_row2, 7, "TOTAL")
        span(header_row1, 8, header_row1, 9, "EMPLOYER'S CONTRIBUTION")
        span(header_row1, 10, header_row2, 10, "TOTAL")
        span(header_row1, 11, header_row2, 11, "REMARKS")

        ws.cell(row=header_row2, column=5, value=f"EPF CONTRIBUTION @ {est.worker_epf_rate:g}% Rs.")
        ws.cell(row=header_row2, column=6, value=f"{est.eps_label} CONTRIBUTION @{est.worker_eps_rate:g}% Rs.")
        ws.cell(row=header_row2, column=8, value=f"EPF CONTRIBUTION @ {est.employer_epf_rate:g}% Rs.")
        ws.cell(row=header_row2, column=9, value=f"{est.eps_label} CONTRIBUTION @{est.employer_eps_rate:g}% Rs.")

        for col in range(1, 12):
            for rr in (header_row1, header_row2):
                c = ws.cell(row=rr, column=col)
                c.font = BOLD
                c.fill = HEADER_FILL
                c.border = BORDER
                c.alignment = CENTER
            c = ws.cell(row=num_row, column=col, value=col)
            c.font = BOLD
            c.border = BORDER
            c.alignment = CENTER

        col_widths_6a = {1: 7, 2: 16, 3: 24, 4: 12, 5: 12, 6: 12, 7: 10, 8: 12, 9: 12, 10: 10, 11: 12}
        ws.row_dimensions[header_row1].height = _row_height_for_cells([
            ("WORKER'S CONTRIBUTION", col_widths_6a[5] + col_widths_6a[6]),
            ("EMPLOYER'S CONTRIBUTION", col_widths_6a[8] + col_widths_6a[9]),
        ])
        ws.row_dimensions[header_row2].height = _row_height_for_cells([
            (ws.cell(row=header_row2, column=5).value, col_widths_6a[5]),
            (ws.cell(row=header_row2, column=6).value, col_widths_6a[6]),
            (ws.cell(row=header_row2, column=8).value, col_widths_6a[8]),
            (ws.cell(row=header_row2, column=9).value, col_widths_6a[9]),
        ])

        row = num_row + 1
        first_data_row = row
        grand = [0, 0, 0, 0, 0, 0, 0]  # wages, w_epf, w_eps, w_tot, e_epf, e_eps, e_tot
        for sl, emp in enumerate(self.employees, start=1):
            values = [sl, format_member_id_display(emp.member_id), emp.name]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = NORMAL
                c.border = BORDER
                c.alignment = CENTER if col_idx == 1 else LEFT
            # Computed directly here (not a live formula reference into the 3A
            # sheet), so the figures always display correctly regardless of
            # the viewer's recalculation settings -- uses the exact same
            # month-by-month rounding as the 3A card, so the two always agree.
            wt, w_epf, w_eps, w_tot, e_epf, e_eps, e_tot = emp.annual_totals(
                est.worker_epf_rate, est.worker_eps_rate, est.employer_epf_rate, est.employer_eps_rate)
            row_values = [wt, w_epf, w_eps, w_tot, e_epf, e_eps, e_tot]
            for col_idx, val in zip(range(4, 11), row_values):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = NORMAL
                c.border = BORDER
                c.alignment = RIGHT
            ws.cell(row=row, column=11, value="").border = BORDER
            for i, val in enumerate(row_values):
                grand[i] += val
            row += 1

        last_data_row = row - 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = BOLD
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = BORDER
        for col_idx, val in zip(range(4, 11), grand):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = BOLD
            c.border = BORDER
            c.alignment = RIGHT
        ws.cell(row=row, column=11).border = BORDER

        widths = {1: 7, 2: 16, 3: 24, 4: 12, 5: 12, 6: 12, 7: 10, 8: 12, 9: 12, 10: 10, 11: 12}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        row += 2
        row = self._write_signature_block(ws, row, num_cols=11)

        self._apply_a4_page_setup(ws, last_row=row, num_cols=11)
        ws.print_title_rows = '8:10'
        footer_text = f"{est.name} | Code: {est.code} | Form - 6A"
        ws.oddFooter.center.text = footer_text
        ws.evenFooter.center.text = footer_text

    # --------------------------------------------------------------- 12A ---
    def _build_12a_sheet(self, wb):
        """
        Form 12A: the monthly consolidated statement of contribution --
        one row per month (not per employee), totalling every employee's
        wages and contributions for that month, plus the statutory
        admin-charge accounts (2, 21, 22) computed on that month's total
        wages. A Grand Total row sums each column down the bottom, and a
        Total column on the right sums every account across for that month.
        """
        est = self.est
        sheet_name = f"12A_{est.short_year_label}" if est.short_year_label else "12A"
        ws = wb.create_sheet(title=sheet_name[:31])
        num_cols = 10

        ws.merge_cells("A1:I1")
        ws["A1"] = "FORM 12 A"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:I2")
        ws["A2"] = "THE EMPLOYEES' PROVIDENT FUND SCHEME, 1952 (PARAGRAPH 38)"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:I3")
        ws["A3"] = (f"Statement of Contribution for the currency period from "
                    f"1st April {est.year_from} to 31st March {est.year_to}")
        ws["A3"].font = BOLD_ITALIC
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:I4")
        ws["A4"] = f"Name & Address of the Establishment :- {est.name}, {est.address}"
        ws["A4"].font = Font(name="Arial", bold=True, italic=True, size=11)
        ws["A4"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A5:K5")
        ws["A5"] = f"Code No. of the Establishment :- {est.code}"
        ws["A5"].font = Font(name="Arial", bold=True, size=11)
        ws["A5"].alignment = Alignment(horizontal="center")

        header_row = 7
        headers = ["Wages Month", "TRRN", "CRRN", "Members", "A/c No.1\n(EE+ER) Rs.",
                   "A/c No.2\n(Admin Chgs.) Rs.", "A/c No.10\n(Pension Fund) Rs.", "A/c No.21\n(EDLI) Rs.",
                   "A/c No.22\n(EDLI Admin) Rs.", "Total\nRs.", "Credit Date"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=i, value=h)
            c.font = BOLD
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = CENTER
            
        num_cols = len(headers)
        col_widths_12a = {1: 12, 2: 18, 3: 18, 4: 10, 5: 14, 6: 14, 7: 15, 8: 12, 9: 14, 10: 13, 11: 13}
        ws.row_dimensions[header_row].height = _row_height_for_cells(
            [(h, col_widths_12a[i]) for i, h in enumerate(headers, start=1)])

        # Every employee's per-month figures, using this year's contribution scheme
        all_month_rows = [emp.month_rows(est.worker_epf_rate, est.worker_eps_rate,
                                          est.employer_epf_rate, est.employer_eps_rate)
        row = header_row + 1
        first_data_row = row
        a2_rates_used, a22_rates_used = [], []
        grand = [0] * 7  # members, a1, a2, a10, a21, a22, total
        
        year_key = f"{est.year_from}-{est.year_to[-2:]}" if est.year_from and est.year_to else ""
        yr_record = self.project.years.get(year_key) if self.project else None
        all_remittances = yr_record.remittances if yr_record and hasattr(yr_record, 'remittances') else []

        for i, month_label in enumerate(MONTHS):
            cal_year = calendar_year_for_month(month_label, est.year_from, est.year_to)
            a2_rate = account2_rate_percent(cal_year, get_month_num(month_label))
            a22_rate = account22_rate_percent(cal_year, get_month_num(month_label))
            a2_rates_used.append((month_label, a2_rate))
            a22_rates_used.append((month_label, a22_rate))

            month_remittances = [r for r in all_remittances if r.get("month_label") == month_label]
            
            if not month_remittances:
                # Fallback to calculated values if no manual remittances are entered for this month
                wages_total = sum(rows[i][0] for rows in all_month_rows)
                ee_total = sum(rows[i][1] for rows in all_month_rows)     # A/c 1 (EE)
                er_total = sum(rows[i][4] for rows in all_month_rows)     # A/c 1 (ER)
                a10_total = sum(rows[i][5] for rows in all_month_rows)    # A/c 10 (Pension Fund)
                
                a2_amt = round(wages_total * a2_rate / 100)
                a21_amt = round(wages_total * ACCOUNT_21_RATE / 100)
                a22_amt = (max(round(wages_total * a22_rate / 100), ACCOUNT_22_MIN)
                          if (a22_rate > 0 and wages_total > 0) else 0)
                
                members = sum(1 for rows in all_month_rows if rows[i][0] > 0)
                acc_01 = ee_total + er_total
                
                # We only show the fallback row if there's actually active wages/members,
                # or if we want to show 0s for every month. The old system showed 0s.
                r_data = {
                    "trrn": "-",
                    "crrn": "-",
                    "members": members,
                    "acc_01": acc_01,
                    "acc_02": a2_amt,
                    "acc_10": a10_total,
                    "acc_21": a21_amt,
                    "acc_22": a22_amt,
                    "credit_date": "-"
                }
                month_remittances.append(r_data)
                
            for r in month_remittances:
                trrn = r.get("trrn", "-")
                crrn = r.get("crrn", "-")
                members = int(r.get("members", 0))
                a1 = int(r.get("acc_01", 0))
                a2 = int(r.get("acc_02", 0))
                a10 = int(r.get("acc_10", 0))
                a21 = int(r.get("acc_21", 0))
                a22 = int(r.get("acc_22", 0))
                cdate = r.get("credit_date", "-")
                row_total = a1 + a2 + a10 + a21 + a22

                row_values = [month_label, trrn, crrn, members, a1, a2, a10, a21, a22, row_total, cdate]
                for col_idx, val in enumerate(row_values, start=1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.font = NORMAL
                    c.border = BORDER
                    c.alignment = CENTER if col_idx <= 4 or col_idx == 11 else RIGHT
                    
                grand[0] += members
                grand[1] += a1
                grand[2] += a2
                grand[3] += a10
                grand[4] += a21
                grand[5] += a22
                grand[6] += row_total
                row += 1

        last_data_row = row - 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = BOLD
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER

        for i_offset, val in enumerate(grand):
            c = ws.cell(row=row, column=4 + i_offset, value=val)
            c.font = BOLD
            c.border = BORDER
            c.alignment = CENTER if i_offset == 0 else RIGHT
            
        c = ws.cell(row=row, column=11, value="-")
        c.font = BOLD
        c.border = BORDER
        c.alignment = CENTER
        
        grand_total_row = row
        row += 2

        # Footnote: exactly which admin-charge rate applied in which month(s),
        # since a financial year can straddle a rate change (e.g. FY 2018-19).
        ws.cell(row=row, column=1,
               value=f"A/c No.2 (Admin Charges) rate applied: {format_rate_periods(a2_rates_used)}")
        ws.cell(row=row, column=1).font = Font(name="Arial", size=8, italic=True)
        row += 1
        ws.cell(row=row, column=1,
               value=f"A/c No.22 (EDLI Admin Charges) rate applied: {format_rate_periods(a22_rates_used)} "
                     f"(Rs. {ACCOUNT_22_MIN} minimum per month where rate > 0%)")
        ws.cell(row=row, column=1).font = Font(name="Arial", size=8, italic=True)
        row += 1

        for col, w in col_widths_12a.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        row += 1
        row = self._write_signature_block(ws, row, num_cols=num_cols)

        self._apply_a4_page_setup(ws, last_row=row, num_cols=num_cols, orientation="landscape",
                                   fit_one_page=True, center_on_page=True)

# --------------------------------------------------------------------------
# PDF export and batch (year-range) generation
# --------------------------------------------------------------------------

def _get_excel_app():
    """
    Starts (or connects to) Excel via COM automation, for PDF export.
    Requires Microsoft Excel installed on this Windows PC, plus the
    'pywin32' package:   pip install pywin32
    Raises RuntimeError with a clear, actionable message if either is missing.
    """
    try:
        import win32com.client as win32
    except ImportError as e:
        raise RuntimeError(
            "PDF export needs Microsoft Excel plus the 'pywin32' package on this "
            "Windows PC. Install it with:\n\n    pip install pywin32\n\n"
            "then try again. (Any Excel files have already been saved.)"
        ) from e
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False   # never pop up a hidden "update links?" dialog
    excel.EnableEvents = False       # skip any workbook_open macros/add-ins that could show a dialog
    return excel


def _export_pdf_with_app(excel_app, xlsx_path: str, pdf_path: str):
    """
    Uses an already-running Excel COM Application object to export one
    workbook to PDF. See convert_workbook_to_pdf for the standalone version.

    If pdf_path already exists (e.g. left open in a PDF viewer from a
    previous run), Excel's ExportAsFixedFormat fails with a vague
    "Document not saved" COM error -- so it's removed first, and if that
    removal itself fails (file locked), a clear message is raised
    immediately instead of letting Excel produce the confusing one.
    """
    xlsx_path = os.path.abspath(xlsx_path)
    pdf_path = os.path.abspath(pdf_path)

    if os.path.exists(pdf_path):
        try:
            os.remove(pdf_path)
        except OSError as e:
            raise RuntimeError(
                f"Could not overwrite the existing PDF:\n{pdf_path}\n\n"
                f"It's probably open in a PDF viewer (Adobe Reader, Edge, a browser tab, "
                f"etc.) -- please close it and try again."
            ) from e

    try:
        wb = excel_app.Workbooks.Open(xlsx_path, UpdateLinks=0, IgnoreReadOnlyRecommended=True)
    except Exception as e:
        raise RuntimeError(
            f"Excel could not open the workbook it just wrote:\n{xlsx_path}\n\n"
            f"({e})\n\n"
            f"If Excel shows a hidden dialog box (e.g. a compatibility or protected-view "
            f"prompt), close it, then also check Task Manager for any leftover background "
            f"EXCEL.EXE process and end it before trying again."
        ) from e

    try:
        try:
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
        except Exception as e:
            raise RuntimeError(
                f"Excel refused to save the PDF for this file:\n{xlsx_path}\n\n"
                f"({e})\n\n"
                f"This is almost always one of:\n"
                f"  - the PDF is already open in a viewer -- close it and retry\n"
                f"  - a leftover invisible EXCEL.EXE process from an earlier run -- end it "
                f"in Task Manager and retry\n"
                f"  - Excel is showing a hidden dialog box off-screen -- check the taskbar\n\n"
                f"The Excel (.xlsx) file itself was already saved successfully; only the "
                f"PDF step failed. You can also untick \"PDF (.pdf)\" and generate Excel-only "
                f"to confirm the data is fine, then retry PDF separately."
            ) from e
    finally:
        wb.Close(SaveChanges=False)
    return pdf_path


def convert_workbook_to_pdf(xlsx_path: str, pdf_path: str):
    """
    Converts a single .xlsx file to .pdf, honouring each sheet's own page
    setup (A4, fit-to-width, print area -- see ExcelGenerator._apply_a4_page_setup).
    Launches and quits its own Excel instance -- fine for a one-off export.
    For generating many years at once, generate_forms_for_year_range shares
    a single Excel instance across the whole batch instead (much faster).
    """
    excel = _get_excel_app()
    try:
        return _export_pdf_with_app(excel, xlsx_path, pdf_path)
    finally:
        excel.Quit()


def generate_forms_for_year(project: "Project", year_key: str, output_dir: str,
                             make_excel: bool = True, make_pdf: bool = True, excel_app=None):
    """
    Generates the Form 3A/6A workbook (and, optionally, a matching PDF) for
    ONE year, saved into output_dir. Both forms live in a single file per
    year, exactly like a single "Generate Excel" click produces.

    excel_app: an already-running Excel COM Application to reuse (see
    generate_forms_for_year_range) -- if None and make_pdf is True, a fresh
    Excel instance is started and closed just for this one call.

    Returns a dict, e.g. {"excel": "...xlsx", "pdf": "...pdf"}.
    """
    os.makedirs(output_dir, exist_ok=True)
    yr = project.years[year_key]
    est = project.build_establishment_for_year(year_key)
    employees = project.build_employees_for_year(year_key)
    gen = ExcelGenerator(est, employees, project=project)

    safe_code = (project.code or "EPF").replace("/", "-").replace("\\", "-").strip() or "EPF"
    base = f"{safe_code}_{yr.short_label}"
    written = {}

    xlsx_path = os.path.join(output_dir, f"{base}.xlsx")
    gen.build(xlsx_path)
    if make_excel:
        written["excel"] = xlsx_path

    if make_pdf:
        pdf_path = os.path.join(output_dir, f"{base}.pdf")
        if excel_app is not None:
            _export_pdf_with_app(excel_app, xlsx_path, pdf_path)
        else:
            convert_workbook_to_pdf(xlsx_path, pdf_path)
        written["pdf"] = pdf_path
        if not make_excel:
            # they only wanted the PDF -- clean up the intermediate xlsx
            try:
                os.remove(xlsx_path)
            except OSError:
                pass

    return written


def generate_forms_for_year_range(project: "Project", start_year_key: str, end_year_key: str,
                                   output_dir: str, make_excel: bool = True, make_pdf: bool = True,
                                   progress_callback=None):
    """
    Generates Form 3A/6A output for every year between start_year_key and
    end_year_key (inclusive), ordered by Year From (e.g. "1997-98" through
    "2014-15"). Each year gets its own file(s) in output_dir.

    If make_pdf, a single Excel instance is started once and reused for
    every year in the range (much faster than launching Excel per year),
    then closed at the end.

    progress_callback(i, total, year_key), if given, is called before each
    year is generated (i is 0-based).

    Returns a list of (year_key, {"excel":..., "pdf":...}) tuples for the
    years that succeeded. Stops and re-raises on the first failure; the
    exception's `.results` attribute holds whatever was completed so far
    (same shape as a normal return value).
    """
    ordered_keys = sorted(project.years.keys(),
                          key=lambda k: natural_sort_key(project.years[k].year_from))
    if start_year_key not in ordered_keys or end_year_key not in ordered_keys:
        raise ValueError("Selected year(s) not found in this project.")
    i_start = ordered_keys.index(start_year_key)
    i_end = ordered_keys.index(end_year_key)
    if i_start > i_end:
        i_start, i_end = i_end, i_start
    keys_in_range = ordered_keys[i_start:i_end + 1]

    excel_app = _get_excel_app() if make_pdf else None
    results = []
    try:
        for i, key in enumerate(keys_in_range):
            if progress_callback:
                progress_callback(i, len(keys_in_range), key)
            try:
                written = generate_forms_for_year(project, key, output_dir, make_excel, make_pdf,
                                                   excel_app=excel_app)
            except Exception as e:
                e.results = results
                raise
            results.append((key, written))
    finally:
        if excel_app is not None:
            excel_app.Quit()
    return results


# --------------------------------------------------------------------------
# Form 5 (new joiners) and Form 10 (leavers), for a single calendar month --
# built straight from the Employee Master's Date of Joining / Date of Exit
# fields. An employee is included in Form 5 ONLY for the exact month their
# Date of Joining falls in, and in Form 10 ONLY for the exact month their
# Date of Exit falls in -- never in every month's list.
# --------------------------------------------------------------------------

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def employees_joined_in_month(project: "Project", cal_year: int, cal_month: int):
    """Employee Master records whose Date of Joining falls in this exact
    calendar month/year, in SL No. order."""
    matches = []
    for m in project.master_list():
        if not m.doj:
            continue
        try:
            d = datetime.strptime(m.doj, "%d/%m/%Y").date()
        except ValueError:
            continue
        if d.year == cal_year and d.month == cal_month:
            matches.append(m)
    return matches


def employees_left_in_month(project: "Project", cal_year: int, cal_month: int):
    """Employee Master records whose Date of Exit falls in this exact
    calendar month/year, in SL No. order."""
    matches = []
    for m in project.master_list():
        if not m.doe:
            continue
        try:
            d = datetime.strptime(m.doe, "%d/%m/%Y").date()
        except ValueError:
            continue
        if d.year == cal_year and d.month == cal_month:
            matches.append(m)
    return matches


NOTE_FONT = Font(name="Arial", size=8)
NAME_ADDR_FONT = Font(name="Arial", bold=True, italic=True, size=11)
CODE_NO_FONT = Font(name="Arial", bold=True, size=11)


def _estimate_wrapped_row_height(text, num_cols, chars_per_col=15, line_height=15, min_height=15):
    """
    openpyxl doesn't auto-fit row height for wrapped text the way Excel's
    UI does -- a merged cell left at the default row height will visually
    clip everything past the first wrapped line. This estimates how many
    lines `text` will wrap to across `num_cols` merged columns (roughly
    chars_per_col characters of usable width each) and returns a row
    height tall enough to show all of them.
    """
    usable_chars = max(20, num_cols * chars_per_col)
    lines_needed = max(1, -(-len(text or "") // usable_chars))  # ceil division
    return max(min_height, lines_needed * line_height + 4)


def _form5_10_header(ws, num_cols, form_title, subtitle_lines, project, intro_line, extra_line=None):
    """
    Header for Form 5 / Form 10, styled to match Form 12A exactly: a bold
    black form title ("FORM 5" / "FORM 10", same TITLE_FONT as "FORM 12 A"),
    centred blue subtitle line(s) combining each scheme with its paragraph
    (SUBTITLE_FONT, like 12A's single "...SCHEME, 1952 (PARAGRAPH 38)"
    line), a bold-italic centred intro sentence ending "...during the month
    of: <month>", an optional extra instruction line (Form 5's "To be sent
    to the Commissioner with Form 2"), and finally single-line "Name &
    Address of the Establishment :- ..." / "Code No. of the Establishment
    :- ..." lines in the same bold-italic / bold fonts 12A uses -- no black
    badge, no separate underlined blank lines. Every line that could wrap
    (the intro sentence, in particular, is long enough to need two lines)
    gets its row height sized to fit, so nothing is clipped. Returns the
    row the column-header row should start on.
    """
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws.cell(row=r, column=1, value=form_title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")
    r += 1

    for text in subtitle_lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        c = ws.cell(row=r, column=1, value=text)
        c.font = SUBTITLE_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[r].height = _estimate_wrapped_row_height(text, num_cols)
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws.cell(row=r, column=1, value=intro_line)
    c.font = BOLD_ITALIC
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = _estimate_wrapped_row_height(intro_line, num_cols)
    r += 1

    if extra_line:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        c = ws.cell(row=r, column=1, value=extra_line)
        c.font = NORMAL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[r].height = _estimate_wrapped_row_height(extra_line, num_cols)
        r += 1

    name_addr_text = f"Name & Address of the Factory/ Establishment :- {project.name}, {project.address}"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws.cell(row=r, column=1, value=name_addr_text)
    c.font = NAME_ADDR_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = _estimate_wrapped_row_height(name_addr_text, num_cols)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws.cell(row=r, column=1, value=f"Code No. of the Factory/ Establishment :- {project.code}")
    c.font = CODE_NO_FONT
    c.alignment = Alignment(horizontal="center")
    r += 2
    return r



def _form5_10_signature_and_notes(ws, row, num_cols, note_paragraphs):
    """
    Leaves a blank spacer row tall enough to actually pen a signature in,
    then writes the ruled signature/seal line (right of centre, matching
    the official form's underline-then-caption layout) followed by a Date
    line on the left, then the statutory note/boilerplate paragraphs
    printed underneath -- matching the bottom of the official Form 5 /
    Form 10. Returns the row after everything written.
    """
    ws.row_dimensions[row].height = 34  # room to actually sign, above the ruled line
    row += 1

    sig_col_start = max(1, num_cols - 4)
    for col in range(sig_col_start, num_cols + 1):
        ws.cell(row=row, column=col).border = Border(top=THIN)
    row += 1
    ws.cell(row=row, column=1, value="Date").font = NORMAL
    ws.merge_cells(start_row=row, start_column=sig_col_start, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=sig_col_start,
                value="Signature of the employer or other authorised officer and "
                      "stamp of the Factory / Establishment")
    c.font = Font(name="Arial", size=9, italic=True)
    c.alignment = CENTER
    row += 2

    for para in note_paragraphs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        c = ws.cell(row=row, column=1, value=para)
        c.font = NOTE_FONT
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 13 * (max(1, len(para) // 130 + 1))
        row += 1
    return row + 1


def _write_form5_sheet(ws, project: "Project", cal_year: int, cal_month: int):
    """
    Writes a complete Form 5 (new joiners, for ONE calendar month), matching
    the official EPFO layout, into an already-created worksheet -- shared by
    the standalone generate_form5_for_month() export and by ExcelGenerator,
    which embeds one Form 5 sheet per month straight into the main
    3A/6A/12A workbook. Returns (last_row, list_of_matched_MasterEmployee).
    """
    matches = employees_joined_in_month(project, cal_year, cal_month)
    month_label = f"{MONTH_NAMES[cal_month - 1]}, {cal_year}"
    num_cols = 10
    col_widths = {1: 6, 2: 16, 3: 16, 4: 22, 5: 30, 6: 12, 7: 8, 8: 20, 9: 36, 10: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    subtitle_lines = [
        "THE EMPLOYEES' PROVIDENT FUND SCHEME, 1952 (PARAGRAPH 36 (2) (a) AND (b))",
        "EMPLOYEES' PENSION SCHEME, 1995 (PARAGRAPH 20 (4))",
    ]
    intro = ("Return of Employees' qualifying for membership of the Employees' Provident Fund, "
             f"Employees' Pension Scheme & Employees' Deposit Linked Insurance Fund for the first "
             f"time during the month of: {month_label}")
    r = _form5_10_header(ws, num_cols, "FORM 5", subtitle_lines, project, intro,
                         extra_line="To be sent to the Commissioner with Form 2")

    header_row = r
    headers = ["S No", "Member ID", "UAN", "Name of the Member",
               "Father's Name or Husband's\nName in case of married women",
               "Age/ Date of\nBirth", "Sex", "Date of Eligibility\nfor Service",
               "Total Period of Previous Service\n(excluding period of breaks) as on\n"
               "the date of joining the fund", "Remarks"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = CENTER
    ws.row_dimensions[header_row].height = _row_height_for_cells(
        [(h, col_widths[i]) for i, h in enumerate(headers, start=1)])

    row = header_row + 1
    num_data_rows = max(len(matches), 10)  # blank numbered rows when nobody joined, like the printed template
    for i in range(1, num_data_rows + 1):
        m = matches[i - 1] if i <= len(matches) else None
        values = ([i, format_member_id_display(m.member_id), m.uan, m.name, m.father_name, m.dob, m.sex, m.doj, "", ""] if m
                  else [i, "", "", "", "", "", "", "", "", ""])
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = CENTER if col_idx in (1, 6) else LEFT
        row += 1

    row += 1
    row = _form5_10_signature_and_notes(ws, row, num_cols, [
        "Note: Please furnish details of the membership in remarks column if the employee was a member of "
        "Employees' Provident Fund and Employees' Family Pension scheme before joining yourself/ factory. "
        "i.e. Member ID and/ or the name and particulars of the last employer.",
    ])

    return row, matches


def generate_form5_for_month(project: "Project", cal_year: int, cal_month: int, filepath: str):
    """
    Standalone export: Form 5 for ONE calendar month, saved as its own
    single-sheet workbook (used by the "Generate Form 5" menu item, which
    lets you produce just one month on demand). Returns
    (filepath, list_of_matched_MasterEmployee).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form 5"
    last_row, matches = _write_form5_sheet(ws, project, cal_year, cal_month)
    ExcelGenerator._apply_a4_page_setup(ws, last_row=last_row, num_cols=9, orientation="landscape",
                                         fit_one_page=True, center_on_page=True)
    wb.save(filepath)
    return filepath, matches


def _write_form10_sheet(ws, project: "Project", cal_year: int, cal_month: int):
    """
    Writes a complete Form 10 (leavers, for ONE calendar month), matching
    the official EPFO layout, into an already-created worksheet -- shared by
    the standalone generate_form10_for_month() export and by ExcelGenerator,
    which embeds one Form 10 sheet per month straight into the main
    3A/6A/12A workbook. Returns (last_row, list_of_matched_MasterEmployee).
    """
    matches = employees_left_in_month(project, cal_year, cal_month)
    month_label = f"{MONTH_NAMES[cal_month - 1]}, {cal_year}"
    num_cols = 8
    col_widths = {1: 6, 2: 16, 3: 16, 4: 22, 5: 28, 6: 17, 7: 16, 8: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    subtitle_lines = [
        "THE EMPLOYEES' PROVIDENT FUND SCHEME, 1952 (PARAGRAPH 36 (2) (a) AND (b))",
        "EMPLOYEES' PENSION SCHEME, 1995 (PARAGRAPH 20 (4))",
    ]
    intro = f"Return of Members leaving service during the month of: {month_label}"
    r = _form5_10_header(ws, num_cols, "FORM 10", subtitle_lines, project, intro)

    header_row = r
    headers = ["S No", "Member ID", "UAN", "Name of the Member",
               "Father's Name or Husband's\nName in case of married",
               "Date of Leaving\nService", "Reason for\nLeaving Service", "Remarks"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = CENTER
    ws.row_dimensions[header_row].height = _row_height_for_cells(
        [(h, col_widths[i]) for i, h in enumerate(headers, start=1)])

    row = header_row + 1
    num_data_rows = max(len(matches), 10)  # blank numbered rows when nobody left, like the printed template
    for i in range(1, num_data_rows + 1):
        m = matches[i - 1] if i <= len(matches) else None
        values = ([i, format_member_id_display(m.member_id), m.uan, m.name, m.father_name, m.doe, m.reason_leaving, ""] if m
                  else [i, "", "", "", "", "", "", ""])
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = CENTER if col_idx in (1, 5) else LEFT
        row += 1

    row += 1
    row = _form5_10_signature_and_notes(ws, row, num_cols, [
        "Please state whether the member is (a) retiring according to para 69 (1) (a) or (b) of the scheme; "
        "(b) leaving India for permanent settlement abroad; (c) retrenched; (d) ordinarily dismissed for serious "
        "and willful misconduct; (e) discharged; (f) resigning from or leaving service; (g) taking up employment "
        "elsewhere (the name and address of the new employer should be stated); (h) dead.",
        "(1) A request for deduction from the account of a member dismissed for serious and willful misconduct "
        "should be reported by the following \"certified that the member mentioned at Sr. No. ___________ Shri "
        "___________ was dismissed from the service for willful misconduct. I recommend that the employer's "
        "contribution for ___________ should be forfeited from his account in the fund. A copy of order of "
        "dismissal is enclosed.",
        "(2) In case of discharge from service, the following certificate should be filled. Certified that the "
        "member mentioned in Sr. No ___________ Shri ___________ was paid/ unpaid retrenchment compensation of "
        "Rs. ___________ under the Industrial Disputes Act, 1947",
    ])

    return row, matches


def generate_form10_for_month(project: "Project", cal_year: int, cal_month: int, filepath: str):
    """
    Standalone export: Form 10 for ONE calendar month, saved as its own
    single-sheet workbook (used by the "Generate Form 10" menu item, which
    lets you produce just one month on demand). Returns
    (filepath, list_of_matched_MasterEmployee).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form 10"
    last_row, matches = _write_form10_sheet(ws, project, cal_year, cal_month)
    ExcelGenerator._apply_a4_page_setup(ws, last_row=last_row, num_cols=7, orientation="landscape",
                                         fit_one_page=True, center_on_page=True)
    wb.save(filepath)
    return filepath, matches


def _write_form9_sheet(ws, project: "Project"):
    """
    Writes Form 9 (Revised) -- the one-time "Return of employees who are
    entitled and required to become members of the Employees' Provident
    Fund and Pension Fund", filed once at the time of coverage, listing
    every employee in the Employee Master (not per-month like Form 5/10).
    Matches the official EPFO layout (all 12 columns, including the
    "Date and reason of leaving service" and office-use settlement-remarks
    columns), styled like Form 12A/5/10. Every column that has a matching
    field in this app's data is populated straight from the Employee
    Master / Establishment record -- Member ID, Name, Father's/Husband's
    Name, Date of Birth, Sex, Date of Joining (as "date of eligibility"),
    and Date of Leaving + Reason all come from project.master_list();
    Establishment Name, Address and Code No. come from the project's
    establishment record. Returns (last_row, list_of_all_MasterEmployee).
    """
    employees = project.master_list()
    num_cols = 13
    # Significantly increased column widths so the natural width exceeds the A4 page. 
    # This forces Excel's fitToWidth=1 to scale it down perfectly to touch the left and right margins.
    col_widths = {1: 8, 2: 18, 3: 18, 4: 26, 5: 26, 6: 14, 7: 10, 8: 14, 9: 16,
                  10: 14, 11: 12, 12: 20, 13: 20}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    subtitle_lines = [
        "THE EMPLOYEES' PROVIDENT FUND SCHEME, 1952 (PARA 36(1))",
        "AND THE EMPLOYEES' PENSION SCHEME, 1995 (PARA 20) (PARA 16(1))",
    ]
    intro = ("Return of employees who are entitled and required to become members of the "
             "Employees' Provident Fund and Pension Fund")
    r = _form5_10_header(ws, num_cols, "FORM 9 (REVISED)", subtitle_lines, project, intro)

    # One-time coverage details the official form asks for. Date of Coverage
    # is now tracked on the Establishment record, so it's filled in
    # automatically; Industry and E.S.I. Code No. aren't tracked anywhere
    # else in this app and are left as blank fill-in lines, like the printed form.
    coverage_line = ("Date of Coverage :- " + project.coverage_date if project.coverage_date
                      else "Date of Coverage :- ______________________")
    for label in (coverage_line,
                  "Industry in which the Factory/ Establishment is engaged :- ______________________",
                  "If covered under the E.S.I. Act, E.S.I. Code No. :- ______________________"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        c = ws.cell(row=r, column=1, value=label)
        c.font = NORMAL
        c.alignment = Alignment(horizontal="center")
        r += 1
    r += 1

    header_row = r
    headers = ["Sr. No", "Member ID", "UAN", "Name of Employee\n(in block capital)",
               "Father's Name (or Husband's\nName in case of married women)",
               "Date of\nBirth", "Sex", "Date of Eligibility\nfor Membership",
               "Total Period of Previous Service\n(excluding period of break) as on\n"
               "date of joining the fund",
               "Machine/ Folio No. of\nLedger Card Opened", "Initials of S.S.",
               "Date and Reason of\nLeaving Service",
               "D.C./S.S./A.A.O./A.C. Remarks\nand Initial on Settlement"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Increase header row height significantly for more vertical expansion
    ws.row_dimensions[header_row].height = 90

    row = header_row + 1
    for i, m in enumerate(employees, start=1):
        leaving = ", ".join(x for x in (m.doe, m.reason_leaving) if x)
        values = [i, format_member_id_display(m.member_id), m.uan, m.name, m.father_name, m.dob, m.sex, m.doj, "", "", "", leaving, ""]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = NORMAL
            c.border = BORDER
            align_horiz = "center" if col_idx in (1, 6, 7, 8) else "left"
            c.alignment = Alignment(horizontal=align_horiz, vertical="center", wrap_text=True)
            
        # Dynamically set row height based on text length, with a taller minimum height
        max_chars = max(len(str(val)) for val in values)
        if max_chars > 40:
            ws.row_dimensions[row].height = 50
        elif max_chars > 20:
            ws.row_dimensions[row].height = 35
        else:
            ws.row_dimensions[row].height = 25
        row += 1
    if not employees:
        for i in range(1, 11):  # blank numbered rows, like the printed template, if the Master is empty
            c = ws.cell(row=row, column=1, value=i)
            c.font = NORMAL
            c.alignment = CENTER
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row, column=col_idx).border = BORDER
            ws.row_dimensions[row].height = 25
            row += 1

    row += 1
    row = _form5_10_signature_and_notes(ws, row, num_cols, [])

    # Repeat table headers
    ws.print_title_rows = f'{header_row}:{header_row}'
    
    # Repeat Establishment Name and Code on Page 2+
    ws.HeaderFooter.differentFirst = True
    header_text = f"Name & Address of the Factory/ Establishment :- {project.name}, {project.address}\nCode No. of the Factory/ Establishment :- {project.code}"
    ws.oddHeader.center.text = header_text
    ws.evenHeader.center.text = header_text

    return row, employees


def generate_form9(project: "Project", filepath: str):
    """
    Standalone export: Form 9 (Revised), the one-time Employee Master
    return, saved as its own single-sheet workbook. Returns
    (filepath, list_of_all_MasterEmployee).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form 9"
    last_row, employees = _write_form9_sheet(ws, project)
    tight_margins = PageMargins(left=0.15, right=0.15, top=0.25, bottom=0.25, header=0.15, footer=0.15)
    ExcelGenerator._apply_a4_page_setup(ws, last_row=last_row, num_cols=13, orientation="landscape",
                                         margins=tight_margins, fit_one_page=False, center_on_page=True)
    wb.save(filepath)
    return filepath, employees

def _build_form5_form10_sheets(self, wb, forms_to_generate=None):
    forms = forms_to_generate or ['5', '10', '9']
    est = self.est
    project = self.project
    tight_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3, header=0.15, footer=0.15)

    if '9' in forms:
        ws9 = wb.create_sheet(title="Form 9")
        last_row9, _ = _write_form9_sheet(ws9, project)
        self._apply_a4_page_setup(ws9, last_row=last_row9, num_cols=13, orientation="landscape",
                                   margins=tight_margins, fit_one_page=False, center_on_page=True)

    for month_abbr in MONTHS:
        cal_year = calendar_year_for_month(month_abbr, est.year_from, est.year_to)
        cal_month = get_month_num(month_abbr)
        if cal_year is None:
            continue

        if '5' in forms:
            matches5 = employees_joined_in_month(project, cal_year, cal_month)
            if matches5:
                ws5 = wb.create_sheet(title=f"F5_{month_abbr}"[:31])
                last_row5, _ = _write_form5_sheet(ws5, project, cal_year, cal_month)
                self._apply_a4_page_setup(ws5, last_row=last_row5, num_cols=9, orientation="landscape",
                                           margins=tight_margins, fit_one_page=True, center_on_page=True)

        if '10' in forms:
            matches10 = employees_left_in_month(project, cal_year, cal_month)
            if matches10:
                ws10 = wb.create_sheet(title=f"F10_{month_abbr}"[:31])
                last_row10, _ = _write_form10_sheet(ws10, project, cal_year, cal_month)
                self._apply_a4_page_setup(ws10, last_row=last_row10, num_cols=7, orientation="landscape",
                                           margins=tight_margins, fit_one_page=True, center_on_page=True)


ExcelGenerator._build_form5_form10_sheets = _build_form5_form10_sheets

def convert_excel_to_pdf(excel_path: str, pdf_path: str):
    import win32com.client
    import os
    import pythoncom
    
    pythoncom.CoInitialize()
    excel = None
    try:
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
        except Exception as e:
            raise RuntimeError(f"Could not start Microsoft Excel. Ensure Excel is installed on the server. ({str(e)})")
            
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        # 0 = xlTypePDF
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
    finally:
        if excel:
            try:
                excel.Quit()
            except:
                pass
        pythoncom.CoUninitialize()

def generate_ecr_month(est, employees: List[Employee], year_record: YearRecord, month_idx: int) -> str:
    """
    Generates ECR text format for a given month index (0-11 for Mar-Feb).
    Format: UAN#~#Name#~#Gross#~#EPF#~#EPS#~#EDLI#~#EE_Share#~#EPS_Share#~#ER_EPF#~#NCP#~#Refund
    """
    lines = []
    wage_ceilings = get_wage_ceilings_for_year(year_record.year_from)
    for emp in employees:
        if not emp.wages or len(emp.wages) <= month_idx:
            continue
        w = emp.wages[month_idx]
        if not w or w <= 0:
            continue

        mrows = emp.month_rows(est.worker_epf_rate, est.worker_eps_rate,
                               est.employer_epf_rate, est.employer_eps_rate,
                               wage_ceilings=wage_ceilings)
        
        if not mrows or len(mrows) <= month_idx:
            continue
            
        r = mrows[month_idx]
        
        w_epf = r[1]
        e_epf = r[4]
        e_eps = r[5]
        
        if emp.gross_wages and len(emp.gross_wages) > month_idx and emp.gross_wages[month_idx]:
            gross = round(emp.gross_wages[month_idx])
        else:
            gross = round(w)
            
        epf_w = round(w)
        eps_w = 0 if emp.age_crosses_58 else round(min(w, wage_ceilings[month_idx]))
        edli_w = round(min(w, wage_ceilings[month_idx]))
        
        # UAN#~#Member Name#~#Gross Wages#~#EPF Wages#~#EPS Wages#~#EDLI Wages#~#EE Share Remitted#~#EPS Contribution Remitted#~#ER EPF Contribution Remitted#~#NCP Days#~#Refund of Advances
        ncp = emp.ncp_days[month_idx] if hasattr(emp, 'ncp_days') and emp.ncp_days and len(emp.ncp_days) > month_idx else 0
        line = f"{emp.uan or ''}#~#{emp.name}#~#{gross}#~#{epf_w}#~#{eps_w}#~#{edli_w}#~#{w_epf}#~#{e_eps}#~#{e_epf}#~#{ncp}#~#0"
        lines.append(line)
    
    return "\n".join(lines)
